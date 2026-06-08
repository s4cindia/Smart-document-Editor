"""Excel reading and writing built on Polars + OpenPyXL + XlsxWriter."""
from __future__ import annotations

from pathlib import Path

import polars as pl


def list_sheets(path: str | Path) -> list[str]:
    """Return worksheet names for an .xlsx/.xls workbook."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    # .xls handled by pandas/xlrd path below
    import pandas as pd
    xls = pd.ExcelFile(path)
    return list(xls.sheet_names)


def read_sheet(path: str | Path, sheet: str | int | None = None) -> pl.DataFrame:
    """Read a single worksheet into a Polars DataFrame."""
    path = Path(path)
    suffix = path.suffix.lower()
    sheet_name = sheet if sheet is not None else 0

    if suffix == ".xlsx":
        # Polars reads xlsx via the calamine/openpyxl engine.
        try:
            return pl.read_excel(path, sheet_name=sheet if isinstance(sheet, str) else None)
        except Exception:  # noqa: BLE001 - fall back to pandas
            pass
    import pandas as pd
    pdf = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    pdf = pdf.where(pd.notnull(pdf), None)
    pdf.columns = [str(c) for c in pdf.columns]
    # Build the Polars frame from plain Python lists rather than via
    # pl.from_pandas(): the latter requires pyarrow for object/nullable
    # ("Int64") columns, which Excel reads produce. strict=False lets Polars
    # infer a sensible dtype per column even when values are mixed.
    data = {col: pdf[col].tolist() for col in pdf.columns}
    return pl.DataFrame(data, strict=False)


def highlight_summary_cells(path: str | Path, sheet_name: str | None = None) -> bool:
    """Post-process an existing .xlsx: fill any "Summary" cell yellow when its
    text is longer than 215 characters or contains a line break.

    Works on a workbook produced by ANY writer (xlsxwriter or openpyxl), so it
    can be applied uniformly to every export that may contain a Summary column.
    Idempotent; only rewrites the file when something actually changed. The
    column is matched case-insensitively from the header row; text is never
    modified or truncated and no other column is touched.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    path = Path(path)
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="FFFF00")
    changed = False
    sheets = ([wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames
              else list(wb.worksheets))
    for ws in sheets:
        summary_cols = [c for c in range(1, ws.max_column + 1)
                        if str(ws.cell(1, c).value or "").strip().lower() == "summary"]
        if not summary_cols:
            continue
        for r in range(2, ws.max_row + 1):
            for c in summary_cols:
                cell = ws.cell(r, c)
                if cell.value is None:
                    continue
                text = str(cell.value)
                if len(text) > 215 or "\n" in text or "\r\n" in text:
                    cell.fill = fill
                    changed = True
    if changed:
        wb.save(path)
    wb.close()
    return changed


def write_excel(df: pl.DataFrame, out_path: str | Path, sheet_name: str = "Sheet1") -> Path:
    """Write a DataFrame to .xlsx WITHOUT highlighting missing values."""
    return _write_xlsx(df, out_path, sheet_name, highlight=False)


def write_excel_highlighted(df: pl.DataFrame, out_path: str | Path,
                            sheet_name: str = "Sheet1") -> Path:
    """Write a DataFrame to .xlsx, highlighting missing/blank cells in blue."""
    return _write_xlsx(df, out_path, sheet_name, highlight=True)


def _write_xlsx(df: pl.DataFrame, out_path: str | Path,
                sheet_name: str = "Sheet1", highlight: bool = False) -> Path:
    """Write a DataFrame to .xlsx in the normal layout: column headers on the
    first row, data below. Every row and column is written.

    A cell counts as "missing" when it is null or an empty/whitespace-only
    string. When *highlight* is True those cells are filled blue; otherwise the
    file is written plainly with no extra formatting changes.
    """
    from utils.helpers import is_missing
    out_path = Path(out_path)
    cols = df.columns
    # Index of the "Summary" column(s), matched case-insensitively. Only these
    # cells get the export-time length / newline highlight.
    summary_idx = {j for j, c in enumerate(cols)
                   if str(c).strip().lower() == "summary"}

    import xlsxwriter
    with xlsxwriter.Workbook(str(out_path)) as wb:
        ws = wb.add_worksheet(sheet_name[:31])
        # No highlighting: header is plain bold text (no fill), and no
        # summary/length highlighting is applied to data cells.
        header_fmt = wb.add_format({"bold": True})
        missing_fmt = wb.add_format({"bg_color": "#2563eb"})

        widths = [len(str(c)) for c in cols]
        for j, col in enumerate(cols):
            ws.write(0, j, col, header_fmt)

        for i, row in enumerate(df.iter_rows(), start=1):
            for j, val in enumerate(row):
                if is_missing(val):
                    if highlight:
                        ws.write_blank(i, j, None, missing_fmt)
                    # plain: leave the cell empty, no formatting
                else:
                    text = str(val)
                    ws.write(i, j, text)
                    if len(text) > widths[j]:
                        widths[j] = len(text)

        for j, w in enumerate(widths):
            ws.set_column(j, j, min(max(w + 2, 8), 60))
        ws.freeze_panes(1, 0)
    return out_path
