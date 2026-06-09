"""Business logic for the four dashboard feature operations.

Reworked to support the workflows shown in the legacy desktop tool while
keeping a clean web/service split:

  * sheet detection + selection            (axe2excel)
  * Parent ID applied to every output row  (axe2excel, downloadable)
  * "Media Inventory" sheet + type filter  (downloadable)
  * preview rows returned to the UI         (all)
  * WCAG / severity / priority breakdowns   (vpat)
  * multi-format export: xlsx, csv, VPAT    (vpat)
  * optional Template_WCAG_Audit.xlsx fill  (axe2excel, if present)

Each public function returns a 3-tuple ``(outputs, stats, preview)`` where
``outputs`` is a list of ``(label, Path)``. Validation problems raise
``ValueError`` with a user-facing message.
"""
from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path

import polars as pl

from config import config
from services import export_service as _ex
from services.excel_service import read_sheet

# Where an optional audit template can be dropped in to enable template output.
TEMPLATE_DIR = config.data_dir / "templates"
WCAG_TEMPLATE = TEMPLATE_DIR / "Template_WCAG_Audit.xlsx"
# Separate template used ONLY by Placeholder 3's "Export using delivery template".
WCAG_DELIVERY_TEMPLATE = TEMPLATE_DIR / "Template_WCAG_Delivery.xlsx"


# --------------------------------------------------------------------------
# Shared helpers
# --------------------------------------------------------------------------
def _read_table(path: str | Path, sheet: str | None = None) -> pl.DataFrame:
    """Read a worksheet (or csv/tsv) into an all-Utf8 Polars frame."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        sep = "\t" if suffix == ".tsv" else ","
        df = pl.read_csv(path, separator=sep, infer_schema_length=0,
                         truncate_ragged_lines=True, ignore_errors=True)
    else:
        df = read_sheet(path, sheet=sheet)
    df = df.rename({c: str(c).strip() for c in df.columns})
    if df.width:
        df = df.with_columns([pl.col(c).cast(pl.Utf8, strict=False) for c in df.columns])
    return df


def list_sheets_for(path: str | Path) -> list[str]:
    """Return worksheet names (xlsx/xls); empty list for csv/tsv."""
    path = Path(path)
    if path.suffix.lower() in (".csv", ".tsv"):
        return []
    try:
        from services.excel_service import list_sheets
        return list_sheets(path)
    except Exception:  # noqa: BLE001
        return []


def _find_col(df: pl.DataFrame, candidates: list[str]) -> str | None:
    lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    for cand in candidates:
        for lc, orig in lower.items():
            if cand.lower() in lc:
                return orig
    return None


def _col_or_blank(df: pl.DataFrame, col: str | None) -> pl.Expr:
    if col and col in df.columns:
        return pl.col(col).cast(pl.Utf8, strict=False).fill_null("")
    return pl.lit("")


def _outpath(stem: str, ext: str = ".xlsx") -> Path:
    return _ex._outpath(config.export_dir, stem, ext)


def _ts() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _slug(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_.\-]+", "_", (text or "").strip())
    return text.strip("_") or "output"


def _preview(df: pl.DataFrame, limit: int = 200) -> dict:
    head = df.head(limit)
    return {
        "columns": head.columns,
        "rows": head.to_dicts(),
        "total": df.height,
        "shown": head.height,
    }


def template_status() -> dict:
    return {"found": WCAG_TEMPLATE.exists(), "path": str(WCAG_TEMPLATE)}


def export_via_template(df: pl.DataFrame, stem: str = "delivery") -> tuple[Path, bool]:
    """Export a frame using the dedicated delivery template if present, else
    plain xlsx. (Placeholder 3 'Export using delivery template'.)"""
    out = _outpath(stem)
    if WCAG_DELIVERY_TEMPLATE.exists():
        try:
            # _fill_template writes after the header and trims that sheet's
            # trailing blanks; the dashboard / component-health sheets are kept.
            _fill_template(df, out, template=WCAG_DELIVERY_TEMPLATE)
            return out, True
        except Exception:  # noqa: BLE001
            pass
    from services.excel_service import write_excel, trim_trailing_blank_rows
    write_excel(df, out, sheet_name="Delivery")
    trim_trailing_blank_rows(out)
    return out, False


def summarize_delivery(df: pl.DataFrame) -> dict:
    """Summarise a loaded audit frame by WCAG Ver and Priority (Show errors)."""
    wcag_c = _find_col(df, ["wcag ver", "wcag version", "wcag sc", "wcag",
                            "success criteria", "criterion", "criteria"])
    prio_c = _find_col(df, ["priority", "prio", "urgency", "severity"])
    out = {"total": df.height, "wcag_col": wcag_c, "prio_col": prio_c,
           "by_wcag": {}, "by_prio": {}}

    def _counts(col):
        g = (df.with_columns(pl.col(col).cast(pl.Utf8, strict=False).fill_null(""))
             .filter(pl.col(col).str.strip_chars() != "")
             .group_by(col).len().sort("len", descending=True))
        return {r[col]: r["len"] for r in g.to_dicts()}

    if wcag_c:
        out["by_wcag"] = _counts(wcag_c)
    if prio_c:
        out["by_prio"] = _counts(prio_c)
    return out


# --------------------------------------------------------------------------
# Feature 1 — Merge axe tool Excel workbooks
# --------------------------------------------------------------------------
def merge_axe(paths: list[str | Path]) -> tuple[list, dict, dict]:
    if not paths:
        raise ValueError("No files were uploaded.")
    frames, per_file = [], []
    for p in paths:
        p = Path(p)
        try:
            df = _read_table(p)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"Could not read '{p.name}': {exc}") from exc
        if df.height == 0 or df.width == 0:
            per_file.append({"file": p.name, "rows": 0, "note": "empty / skipped"})
            continue
        df = df.with_columns(pl.lit(p.name).alias("Source File"))
        frames.append(df)
        per_file.append({"file": p.name, "rows": df.height})
    if not frames:
        raise ValueError("None of the uploaded files contained any data rows.")

    merged = pl.concat(frames, how="diagonal_relaxed")
    cols = [c for c in merged.columns if c != "Source File"] + ["Source File"]
    merged = merged.select(cols)
    before = merged.height
    content_cols = [c for c in merged.columns if c != "Source File"]
    merged = merged.unique(subset=content_cols, keep="first", maintain_order=True)
    duplicates = before - merged.height

    out = _outpath(f"merged_axe_{_ts()}")
    from services.excel_service import write_excel, highlight_summary_cells
    write_excel(merged, out, sheet_name="Merged")
    # highlighting removed per request: highlight_summary_cells(out)

    stats = {
        "files_processed": len([f for f in per_file if f.get("rows")]),
        "files_total": len(paths),
        "total_rows_in": before,
        "duplicates_removed": duplicates,
        "final_rows": merged.height,
        "columns": merged.width,
        "per_file": per_file,
    }
    return [("Merged workbook", out)], stats, _preview(merged)


# --------------------------------------------------------------------------
# Feature 2 — Generate Axe 2 Excel  (faithful port of generate_axe_excel.py)
#
# Reads a (merged) axe DevTools export workbook by FIXED COLUMN LETTERS and
# fills a copy of the WCAG audit template's "2 - All Issues" sheet:
#   * straight column copies (COLUMN_MAP)
#   * first WCAG criterion id  -> N, name -> O, level -> F (wcag_tags.txt)
#   * severity -> priority (B) and urgency (U)
#   * keyword-derived issue category -> C (validated against rule id)
#   * composed "How to find / fix" detail -> T, element+selector -> H, etc.
# --------------------------------------------------------------------------
TARGET_SHEET = "2 - All Issues"
WCAG_TAGS = TEMPLATE_DIR / "wcag_tags.txt"
DEFAULT_INPUT_START_ROW = 2
DEFAULT_OUTPUT_START_ROW = 2

COLUMN_MAP = [("B", "D"), ("O", "E"), ("A", "G"), ("R", "I"), ("J", "J")]

SEVERITY_COLUMN = "E"
SEVERITY_PRIORITY_COLUMN = "B"
SEVERITY_URGENCY_COLUMN = "U"
SEVERITY_MAP = {
    "critical": ("P1 - Immediate", "1 - Urgent"),
    "serious":  ("P2 - High",      "2 - High"),
    "moderate": ("P3 - Medium",    "3 - Medium"),
    "minor":    ("P4 - Low",       "4 - Low"),
}

WCAG_TAG_COLUMN = "L"
WCAG_ID_COLUMN = "N"
WCAG_NAME_COLUMN = "O"
WCAG_LEVEL_COLUMN = "F"
ISSUE_DESC_COLUMN = "B"
ISSUE_CATEGORY_COLUMN = "C"

_WCAG_TAG_RE = re.compile(r"^wcag(\d{3,4})(a?)$", re.IGNORECASE)
_WCAGAA_TAG_RE = re.compile(r"^wcag(\d{1,2}aa?)$", re.IGNORECASE)

# Header labels used when no template is installed (standalone fallback).
_STANDALONE_HEADERS = {
    "A": "Issue ID", "B": "Priority", "C": "Issue Category", "D": "Description",
    "E": "Component / Rule", "F": "Level / Standard", "G": "Rule ID",
    "H": "Element / Selector", "I": "URL / Screenshot", "J": "Resolution",
    "K": "Fix Owner", "L": "Reopen Count", "M": "Comment Count",
    "N": "WCAG SC", "O": "WCAG Name", "P": "Status",
    "Q": "Project", "R": "Type", "S": "Summary", "T": "Details", "U": "Urgency",
    "V": "Owner Email", "W": "Parent ID",
}

# column widths + which columns wrap long text (standalone formatting)
_STANDALONE_WIDTHS = {
    "A": 12, "B": 15, "C": 22, "D": 48, "E": 20, "F": 16, "G": 18, "H": 36,
    "I": 30, "J": 40, "K": 12, "L": 12, "M": 13, "N": 10, "O": 26, "P": 12,
    "Q": 10, "R": 12, "S": 50, "T": 64, "U": 14, "V": 30, "W": 14,
}
_STANDALONE_WRAP = {"C", "D", "H", "I", "J", "O", "S", "T"}

_KEYWORD_RULES = [
    ("contrast", "Color Contrast"),
    ("focus indicator", "CSS (Focus)"), ("focus visible", "CSS (Focus)"),
    ("focus-visible", "CSS (Focus)"), ("focus outline", "CSS (Focus)"),
    ("keyboard focus", "CSS (Focus)"), (":focus", "CSS (Focus)"),
    ("keyboard alone", "Keyboard"), ("keyboard operable", "Keyboard"),
    ("keyboard accessible", "Keyboard"), ("by keyboard", "Keyboard"),
    ("keyboard", "Keyboard"),
    ("status message", "Status Messages"),
    ("automatically announced", "Status Messages"),
    ("live region", "Status Messages"),
    ("landmark", "Landmarks"), ("region", "Landmarks"),
    ("heading", "Headings"),
    ("list item", "Lists"), ("<li>", "Lists"),
    ("iframe", "Frames / Iframes"), ("frame", "Frames / Iframes"),
    ("label", "Forms / Labels"), ("form field", "Forms / Labels"),
    ("input field", "Forms / Labels"), ("autocomplete", "Forms / Labels"),
    ("touch target", "Target Size"), ("target size", "Target Size"),
    ("minimum size", "Target Size"), ("size or spacing", "Target Size"),
    ("size and space", "Target Size"),
    ("alt text", "Image"), ("alternative text", "Image"),
    ("text alternative", "Image"), ("alt attribute", "Image"),
    ("convey", "Use of Color / State Indication"),
    ("color alone", "Use of Color / State Indication"),
    ("button", "Buttons"),
    ("accessible name", "ARIA / Widget"), ("aria", "ARIA / Widget"),
    ("role", "ARIA / Widget"), ("interactive control", "ARIA / Widget"),
]

_RULEID_CATEGORY_HINTS = {
    "color-contrast": {"Color Contrast"},
    "link-in-text-block": {"Color Contrast", "Use of Color / State Indication"},
    "focus": {"CSS (Focus)", "Keyboard"},
    "landmark": {"Landmarks"}, "region": {"Landmarks"},
    "heading": {"Headings"}, "empty-heading": {"Headings"},
    "page-has-heading-one": {"Headings"},
    "listitem": {"Lists"}, "definition-list": {"Lists"}, "list": {"Lists"},
    "frame-title": {"Frames / Iframes"}, "frame-tested": {"Frames / Iframes"},
    "frame": {"Frames / Iframes"},
    "label-title-only": {"Forms / Labels"},
    "label-content-name-mismatch": {"Forms / Labels", "ARIA / Widget"},
    "form-field-multiple-labels": {"Forms / Labels"},
    "select-name": {"Forms / Labels"}, "autocomplete-valid": {"Forms / Labels"},
    "label": {"Forms / Labels"},
    "image-alt": {"Image"}, "input-image-alt": {"Image"},
    "role-img-alt": {"Image"}, "svg-img-alt": {"Image"}, "area-alt": {"Image"},
    "image-redundant-alt": {"Image"}, "object-alt": {"Image"},
    "button-name": {"Buttons", "ARIA / Widget"},
    "input-button-name": {"Buttons", "ARIA / Widget"},
    "link-name": {"Buttons", "ARIA / Widget"},
    "target-size": {"Target Size"},
    "aria-live": {"Status Messages", "ARIA / Widget"},
    "nested-interactive": {"Keyboard", "ARIA / Widget"},
    "scrollable-region-focusable": {"Keyboard"}, "keyboard": {"Keyboard"},
    "aria-": {"ARIA / Widget", "Forms / Labels", "Status Messages", "Buttons", "Image"},
    "aria": {"ARIA / Widget"}, "role": {"ARIA / Widget"},
}


def _wcag_ids_from_cell(value) -> list[str]:
    if value is None:
        return []
    seen = []
    for token in str(value).split(","):
        m = _WCAG_TAG_RE.match(token.strip().lower())
        if m:
            digits, suffix = m.group(1), m.group(2)
            criterion = f"{digits[0]}.{digits[1]}.{digits[2:]}"
            if suffix:
                criterion += "." + suffix
            if criterion not in seen:
                seen.append(criterion)
    return seen


def _category_allowed_for_ruleid(category, ruleid) -> bool:
    if not ruleid:
        return True
    rid = str(ruleid).strip().lower()
    if not rid:
        return True
    for key in sorted(_RULEID_CATEGORY_HINTS, key=len, reverse=True):
        if key in rid:
            return category in _RULEID_CATEGORY_HINTS[key]
    return True


def _category_from_keywords(description, ruleid=None):
    if not description:
        return None
    text = str(description).lower()
    category = None
    for needle, cat in _KEYWORD_RULES:
        if needle in text:
            category = cat
            break
    if category is None:
        return None
    if not _category_allowed_for_ruleid(category, ruleid):
        return None
    return category


def load_wcag_tags(path: Path) -> dict:
    mapping = {}
    if not path or not Path(path).exists():
        return mapping
    with open(path, encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3 and parts[0]:
                mapping[parts[0]] = (parts[1], parts[2])
    return mapping


def _input_worksheet(path: Path, sheet: str | None):
    """Return (worksheet, max_row, close_fn). Supports xlsx/xls and csv/tsv."""
    from openpyxl import Workbook, load_workbook
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        sep = "\t" if suffix == ".tsv" else ","
        df = pl.read_csv(path, separator=sep, infer_schema_length=0,
                         truncate_ragged_lines=True, ignore_errors=True)
        wb = Workbook()
        ws = wb.active
        ws.append(list(df.columns))
        for row in df.iter_rows():
            ws.append(list(row))
        return ws, ws.max_row, wb.close
    wb = load_workbook(path, data_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "Merged" in wb.sheetnames:
        ws = wb["Merged"]
    else:
        ws = wb.active
    return ws, ws.max_row, wb.close


def axe_to_audit(path: str | Path, sheet: str | None = None,
                 parent_id: str = "", out_name: str = "") -> tuple[list, dict, dict]:
    """Faithful web port of generate_axe_excel.fill_template()."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    path = Path(path)
    in_ws, in_max, close_src = _input_worksheet(path, sheet)

    tags_found = WCAG_TAGS.exists()
    tags_map = load_wcag_tags(WCAG_TAGS)

    # output workbook: copy of the template, or a standalone fallback
    template_used = False
    if WCAG_TEMPLATE.exists():
        tpl = load_workbook(WCAG_TEMPLATE)
        if TARGET_SHEET in tpl.sheetnames:
            out_ws = tpl[TARGET_SHEET]
        else:
            out_ws = tpl.active
        out_wb = tpl
        template_used = True
    else:
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = TARGET_SHEET
        for col, label in _STANDALONE_HEADERS.items():
            out_ws[f"{col}1"] = label

    in_start = DEFAULT_INPUT_START_ROW
    out_row = DEFAULT_OUTPUT_START_ROW
    written = wcag_rows = matched = severity_rows = 0
    missing: set[str] = set()

    def _cv(col, row):
        return in_ws[f"{col}{row}"].value

    for in_row in range(in_start, in_max + 1):
        cells = {}
        first_id = name = level = ""
        non_empty = False

        # 1. straight column copies
        for in_col, out_col in COLUMN_MAP:
            v = _cv(in_col, in_row)
            cells[out_col] = v
            if v is not None and str(v).strip() != "":
                non_empty = True

        # 2/3. WCAG id (first only) + name/level lookup
        ids = _wcag_ids_from_cell(_cv(WCAG_TAG_COLUMN, in_row))
        if ids:
            first_id = ids[0]
            if first_id.endswith(".a"):
                first_id = first_id[:-2]
            cells[WCAG_ID_COLUMN] = first_id
            non_empty = True
            wcag_rows += 1
            if first_id in tags_map:
                name, level = tags_map[first_id]
                cells[WCAG_NAME_COLUMN] = f"{name}"
                cells[WCAG_LEVEL_COLUMN] = level
                matched += 1
            else:
                missing.add(first_id)

        # 4. severity -> priority / urgency
        sev_raw = _cv(SEVERITY_COLUMN, in_row)
        if sev_raw is not None:
            sev = str(sev_raw).strip().lower()
            if sev in SEVERITY_MAP:
                priority, urgency = SEVERITY_MAP[sev]
                cells[SEVERITY_PRIORITY_COLUMN] = priority
                cells[SEVERITY_URGENCY_COLUMN] = urgency
                non_empty = True
                severity_rows += 1

        # 4b. issue category from description, validated against rule id (A)
        desc_val = _cv(ISSUE_DESC_COLUMN, in_row)
        ruleid_val = _cv("A", in_row)
        issue_category = _category_from_keywords(desc_val, ruleid_val)
        if issue_category:
            cells[ISSUE_CATEGORY_COLUMN] = issue_category
            non_empty = True

        root_cause = _cv("B", in_row)
        component = _cv("O", in_row)
        component = component.split("?")[0] if component else ""
        element = str(_cv("K", in_row))
        selector = _cv("I", in_row)
        issue_url_screenshot = _cv("R", in_row)
        help_url = _cv("D", in_row)

        howtofind = howtofix = ""
        if help_url is not None and str(help_url).strip() != "":
            howtofind = (f"- HOW TO FIND - \nREPRODUCE:\n  Issue Type:\n"
                         f"{issue_category}\nComponent: {component}\n  URL: "
                         f"{issue_url_screenshot}\n  Element: {element}\n  "
                         f"Selector: {selector}\nWHERE TO FIX: Review the "
                         f"source/component shown above.\nVERIFY: Re-scan after "
                         f"the fix is deployed.\n\n")
            howtofix = (f"- HOW TO FIX - \nRule ID: {ruleid_val}\nImpact: "
                        f"{str(_cv('E', in_row))}\nIssue: {root_cause}\nHelp: "
                        f"{str(_cv('C', in_row))}\nRule URL: {help_url}\nFix:  "
                        f"apply the appropriate accessibility correction for "
                        f"{ruleid_val}")
        t_desc = (f"\n\nFix Owner:\nVendor\n\nWCAG Ref:\nRef:{first_id}\n\n"
                  f"WCAG:\n{first_id} {name}")
        cells["T"] = howtofind + howtofix + t_desc
        cells["H"] = str(element) + " " + str(selector)

        # fixed summary columns
        cells["K"] = "Vendor"
        cells["L"] = "0"
        cells["M"] = "0"
        cells["P"] = "Pending"
        cells["Q"] = "PSGIN"
        cells["R"] = "Story"
        cells["V"] = "kevin.murphy@ascendlearning.com"

        # standard tag (e.g. WCAG2A / WCAG21AA) -> F (overrides level)
        standard_id = []
        for token in str(_cv(WCAG_TAG_COLUMN, in_row)).split(","):
            if _WCAGAA_TAG_RE.match(token.strip().lower()):
                standard_id.append(token.strip().upper())
        if standard_id:
            cells["F"] = standard_id[0]

        if parent_id:
            cells["W"] = parent_id

        if not non_empty:
            continue

        cells["S"] = (f"Issue - {(out_row - 1):02d} - {str(cells.get('C', ''))} "
                      f"- {component} - {str(cells.get('G', ''))}")
        cells["A"] = f"Issue - {(out_row - 1):02d}"

        for out_col, value in cells.items():
            cell = out_ws[f"{out_col}{out_row}"]
            cell.value = value
            if out_col == "T":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        out_row += 1
        written += 1

    close_src()

    stem = _slug(out_name) if out_name else f"{_slug(path.stem)}_output"
    out = _outpath(stem)
    if not template_used:
        _format_standalone_sheet(out_ws, out_row - 1)
    out_wb.save(out)
    from services.excel_service import trim_trailing_blank_rows
    trim_trailing_blank_rows(out)
    from services.excel_service import highlight_summary_cells
    # highlighting removed per request: highlight_summary_cells(out)

    warnings = []
    if not template_used:
        warnings.append("No template at data/templates/Template_WCAG_Audit.xlsx "
                        "— generated a standalone workbook with the same columns.")
    if not tags_found:
        warnings.append(f"wcag_tags.txt not found — columns "
                        f"{WCAG_NAME_COLUMN}/{WCAG_LEVEL_COLUMN} left blank where "
                        f"no standard tag applied.")
    elif missing:
        shown = ", ".join(sorted(missing)[:12])
        more = "" if len(missing) <= 12 else f" (+{len(missing) - 12} more)"
        warnings.append(f"{len(missing)} id(s) not found in wcag_tags.txt: "
                        f"{shown}{more}")

    stats = {
        "written": written,
        "wcag_rows": wcag_rows,
        "matched": matched,
        "severity_rows": severity_rows,
        "sheet": sheet or "(auto)",
        "parent_id": parent_id or "",
        "template_used": template_used,
        "tags_found": tags_found,
        "target_sheet": TARGET_SHEET,
        "warnings": warnings,
    }
    return [("Audit workbook", out)], stats, _read_output_preview(out, out_ws.title)


def _format_standalone_sheet(ws, last_row: int) -> None:
    """Make the fallback workbook a clean, readable document."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hdr_fill = None  # no header highlighting
    hdr_font = Font(bold=True, size=11)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=False)
    thin = Side(style="thin", color="E4E8F0")
    border = Border(bottom=thin, right=thin)

    for col, width in _STANDALONE_WIDTHS.items():
        ws.column_dimensions[col].width = width
    for col in _STANDALONE_HEADERS:
        cell = ws[f"{col}1"]
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[1].height = 26

    for row in range(2, max(last_row, 1) + 1):
        for col in _STANDALONE_HEADERS:
            cell = ws[f"{col}{row}"]
            cell.border = border
            cell.alignment = (Alignment(vertical="top", wrap_text=True)
                              if col in _STANDALONE_WRAP else body_align)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:W{max(last_row, 1)}"


def _read_output_preview(out_path: Path, sheet_name: str, limit: int = 200) -> dict:
    """Read back the populated sheet for the UI preview table."""
    from openpyxl import load_workbook
    wb = load_workbook(out_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or ()
    max_c = len(header)
    data_rows = []
    for r in rows_iter:
        data_rows.append(r)
        if len(data_rows) >= limit:
            break
    wb.close()
    # keep only columns that have a header or any data
    keep = []
    for i in range(max_c):
        label = header[i]
        has_data = any(i < len(r) and r[i] not in (None, "") for r in data_rows)
        if (label not in (None, "")) or has_data:
            keep.append(i)
    from openpyxl.utils import get_column_letter
    columns = [str(header[i]) if header[i] not in (None, "")
               else get_column_letter(i + 1) for i in keep]
    rows = []
    for r in data_rows:
        rows.append({columns[j]: ("" if (keep[j] >= len(r) or r[keep[j]] is None)
                                   else str(r[keep[j]])) for j in range(len(keep))})
    # total data rows
    total = max(0, (ws.max_row or 1) - 1) if hasattr(ws, "max_row") else len(rows)
    return {"columns": columns, "rows": rows, "total": len(rows), "shown": len(rows)}


def _fill_template(audit: pl.DataFrame, out: Path, template: Path = WCAG_TEMPLATE) -> None:
    """Fill a template's '2 - All Issues' sheet, writing data right after the
    header row (overwriting any pre-formatted blank rows), then drop leftover
    trailing blank rows on that sheet so there is no gap and no empty tail."""
    from openpyxl import load_workbook
    wb = load_workbook(template)
    ws = None
    for name in wb.sheetnames:
        if "all issues" in name.lower() or name.strip().startswith("2"):
            ws = wb[name]
            break
    ws = ws or wb.active
    # Header = first non-empty row (normally row 1); data begins immediately after.
    header_row = 1
    for r in range(1, (ws.max_row or 1) + 1):
        if any(ws.cell(r, c).value not in (None, "")
               for c in range(1, (ws.max_column or 1) + 1)):
            header_row = r
            break
    start = header_row + 1
    for i, row in enumerate(audit.iter_rows(), start=start):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    wb.save(out)
    # Remove any pre-formatted blank rows left below the data on the issues sheet.
    from services.excel_service import trim_trailing_blank_rows
    trim_trailing_blank_rows(out, sheet_name=ws.title)
# Feature 3 — Generate audit Excel for downloadable documents
# --------------------------------------------------------------------------
_DOC_TYPES = {".pdf": "PDF", ".doc": "Microsoft Word", ".docx": "Microsoft Word",
              ".ppt": "Microsoft PowerPoint", ".pptx": "Microsoft PowerPoint"}
_TYPE_KEYWORDS = [("pdf", "PDF"),
                  ("powerpoint", "Microsoft PowerPoint"),
                  ("pptx", "Microsoft PowerPoint"), ("ppt", "Microsoft PowerPoint"),
                  ("word", "Microsoft Word"), ("docx", "Microsoft Word"),
                  ("doc", "Microsoft Word")]


def _classify_doc(*values: str) -> str | None:
    blob = " ".join(v.lower() for v in values if v)
    for ext, friendly in _DOC_TYPES.items():
        if ext in blob:
            return friendly
    for kw, friendly in _TYPE_KEYWORDS:
        if kw in blob:
            return friendly
    return None


def downloadable_docs(path: str | Path, parent_id: str = "",
                      sheet: str | None = None) -> tuple[list, dict, dict]:
    # prefer a "Media Inventory" sheet when present (legacy behaviour)
    if sheet is None:
        for s in list_sheets_for(path):
            if "media" in s.lower() and "inventory" in s.lower():
                sheet = s
                break
    df = _read_table(path, sheet=sheet)
    if df.height == 0:
        raise ValueError("The media inventory sheet has no data rows.")

    url_c = _find_col(df, ["url", "location", "link", "href", "file", "filename",
                           "path", "document", "resource", "media url", "source"])
    name_c = _find_col(df, ["title", "name", "document name", "file name",
                            "filename", "asset"])
    type_c = _find_col(df, ["document type", "type", "format", "file type",
                            "extension", "mime", "kind", "category"])
    notes_c = _find_col(df, ["notes", "note", "comment", "remarks"])

    if url_c is None and type_c is None and name_c is None:
        raise ValueError(
            "Could not find a Title, Type or URL/Location column in the "
            "inventory. Found: " + ", ".join(df.columns[:12])
            + ("…" if df.width > 12 else "")
        )

    counts = {"PDF": 0, "Microsoft Word": 0, "Microsoft PowerPoint": 0}
    matched = []
    for r in df.to_dicts():
        url = str(r.get(url_c, "") or "") if url_c else ""
        name = str(r.get(name_c, "") or "") if name_c else ""
        typ = str(r.get(type_c, "") or "") if type_c else ""
        notes = str(r.get(notes_c, "") or "") if notes_c else ""
        friendly = _classify_doc(typ, url, name)
        if friendly:
            counts[friendly] = counts.get(friendly, 0) + 1
            matched.append({
                "Title": name or (url.rsplit("/", 1)[-1] if url else ""),
                "Document Type": friendly,
                "URL / Location": url,
                "Notes": notes,
                "Review Status": "Pending Review",
                "Parent ID": parent_id or "",
            })
    if not matched:
        raise ValueError("No PDF, Word or PowerPoint documents were found in "
                         "the inventory.")

    out_df = pl.DataFrame(
        [{"S.No": i + 1, **m} for i, m in enumerate(matched)]
    ).select(["S.No", "Title", "Document Type", "URL / Location", "Notes",
              "Review Status", "Parent ID"])

    stem = f"{_slug(parent_id)}_downloadable" if parent_id else f"downloadable_docs_{_ts()}"
    out = _outpath(stem)
    from services.excel_service import write_excel
    write_excel(out_df, out, sheet_name="Downloadable Docs")

    stats = {
        "inventory_rows": df.height,
        "matched": len(matched),
        "sheet": sheet or "(first sheet)",
        "parent_id": parent_id or "",
        "by_type": {k: v for k, v in counts.items() if v},
        "counts_all": {"PDF": counts.get("PDF", 0),
                       "Microsoft Word": counts.get("Microsoft Word", 0),
                       "Microsoft PowerPoint": counts.get("Microsoft PowerPoint", 0)},
    }
    return [("Downloadable documents", out)], stats, _preview(out_df)


# --------------------------------------------------------------------------
# Feature 4 — VPAT / delivery outputs
# --------------------------------------------------------------------------
def vpat_report(path: str | Path) -> tuple[list, dict, dict]:
    df = _read_table(path)
    if df.height == 0:
        raise ValueError("The uploaded audit workbook has no data rows.")

    desc_c = _find_col(df, ["issue description", "description", "issue", "help",
                            "message", "summary", "violation", "finding"])
    wcag_c = _find_col(df, ["wcag sc", "wcag", "success criteria", "criterion",
                            "criteria", "guideline", "standard", "sc"])
    impact_c = _find_col(df, ["impact", "severity", "priority", "level"])
    prio_c = _find_col(df, ["priority", "urgency", "p1", "rank"])

    if desc_c is None and wcag_c is None:
        raise ValueError(
            "Could not recognise audit columns. Expected an issue/description "
            "column and ideally a WCAG criterion column. Found: "
            + ", ".join(df.columns[:12]) + ("…" if df.width > 12 else "")
        )

    work = df.select([
        _col_or_blank(df, desc_c).alias("Issue Description"),
        _col_or_blank(df, wcag_c).alias("WCAG SC"),
        _col_or_blank(df, impact_c).alias("Severity"),
        _col_or_blank(df, prio_c).alias("Priority"),
    ]).filter(
        (pl.col("Issue Description").str.strip_chars() != "")
        | (pl.col("WCAG SC").str.strip_chars() != "")
    )
    total = work.height
    if total == 0:
        raise ValueError("No accessibility findings were detected in the workbook.")

    unique_issues = work.unique(subset=["WCAG SC", "Issue Description"],
                                maintain_order=True)
    n_unique = unique_issues.height

    def _breakdown(col: str) -> dict:
        g = (work.filter(pl.col(col).str.strip_chars() != "")
             .group_by(col).len().sort("len", descending=True))
        return {row[col]: row["len"] for row in g.to_dicts()}

    by_wcag = _breakdown("WCAG SC")
    by_sev = _breakdown("Severity")
    by_prio = _breakdown("Priority")

    # VPAT conformance table (unique issues grouped by criterion)
    crit = (unique_issues.filter(pl.col("WCAG SC").str.strip_chars() != "")
            .group_by("WCAG SC")
            .agg(pl.len().alias("count"),
                 pl.col("Issue Description").first().alias("sample"))
            .sort("count", descending=True))
    vpat_rows = []
    for row in crit.to_dicts():
        n = row["count"]
        vpat_rows.append({
            "Criteria": row["WCAG SC"],
            "Conformance Level": "Does Not Support" if n >= 3 else "Partially Supports",
            "Issue Count": n,
            "Remarks and Explanations":
                f"{n} issue(s) identified. e.g. {str(row['sample'])[:160]}",
        })
    if not vpat_rows:
        for k, v in (by_sev or by_prio).items():
            vpat_rows.append({"Criteria": k, "Conformance Level": "Partially Supports",
                              "Issue Count": v,
                              "Remarks and Explanations": f"{v} issue(s)."})

    # build outputs: findings workbook (xlsx) + csv, plus the VPAT report
    from services.excel_service import write_excel
    findings = work.with_columns(pl.Series("S.No", list(range(1, total + 1)))) \
                   .select(["S.No", "Issue Description", "WCAG SC", "Severity", "Priority"])
    xlsx_out = _outpath("vpat_findings")
    write_excel(findings, xlsx_out, sheet_name="Findings")
    csv_out = _outpath("vpat_findings", ".csv")
    findings.write_csv(csv_out)
    # 3) VPAT report workbook
    summary_rows = [
        {"Metric": "Total findings", "Value": total},
        {"Metric": "Unique issues", "Value": n_unique},
        {"Metric": "WCAG criteria affected", "Value": len(by_wcag)},
    ] + [{"Metric": f"Severity – {k}", "Value": v} for k, v in by_sev.items()] \
      + [{"Metric": f"Priority – {k}", "Value": v} for k, v in by_prio.items()]
    vpat_out = _outpath("vpat_report")
    _write_vpat_workbook(vpat_out, summary_rows, vpat_rows)

    stats = {
        "total_findings": total,
        "unique_issues": n_unique,
        "criteria_affected": len(by_wcag),
        "by_wcag": by_wcag,
        "by_severity": by_sev,
        "by_priority": by_prio,
    }
    outputs = [("Delivery file (Excel)", xlsx_out),
               ("Delivery file (CSV)", csv_out),
               ("Summary report", vpat_out)]
    return outputs, stats, _preview(findings)


def _write_vpat_workbook(out: Path, summary_rows: list[dict],
                         vpat_rows: list[dict]) -> Path:
    import xlsxwriter
    with xlsxwriter.Workbook(str(out)) as wb:
        title = wb.add_format({"bold": True, "font_size": 14, "font_color": "#1d4ed8"})
        hdr = wb.add_format({"bold": True, "border": 1})
        cell = wb.add_format({"border": 1, "valign": "top", "text_wrap": True})
        s = wb.add_worksheet("Summary")
        s.merge_range(0, 0, 0, 1, "Product Accessibility Report — Summary", title)
        s.write(2, 0, "Metric", hdr); s.write(2, 1, "Value", hdr)
        for i, row in enumerate(summary_rows, start=3):
            s.write(i, 0, row["Metric"], cell); s.write(i, 1, row["Value"], cell)
        s.set_column(0, 0, 34); s.set_column(1, 1, 16)
        v = wb.add_worksheet("VPAT")
        cols = ["Criteria", "Conformance Level", "Issue Count", "Remarks and Explanations"]
        v.merge_range(0, 0, 0, len(cols) - 1,
                      "WCAG 2.x Report (auto-generated — review before publishing)", title)
        for j, c in enumerate(cols):
            v.write(2, j, c, hdr)
        for i, row in enumerate(vpat_rows, start=3):
            for j, c in enumerate(cols):
                v.write(i, j, row.get(c, ""), cell)
        v.set_column(0, 0, 26); v.set_column(1, 1, 20)
        v.set_column(2, 2, 12); v.set_column(3, 3, 70)
        v.freeze_panes(3, 0)
    return out
