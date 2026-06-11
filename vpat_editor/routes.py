"""VPAT Editor routes — page + PDF export + draft save/open."""
from __future__ import annotations

import datetime
import json
import logging
import re
from io import BytesIO

from flask import (Blueprint, jsonify, make_response, render_template, request, send_file)

from config import config
from utils import file_utils
from vpat_editor.vpat_core import WCAG_SLUGS, default_data, export_pdf, sc_url

log = logging.getLogger("sde.vpat_editor")

VPAT_BUILD = "43"

vpat_editor_bp = Blueprint("vpat_editor", __name__, template_folder="templates")


# --------------------------------------------------------------------------
# Excel analysis helpers (Browse → Analyse)
# --------------------------------------------------------------------------
_CONF_RANK = {"Does Not Support": 4, "Partially Supports": 3, "Supports": 2,
              "Not Applicable": 1, "Not Evaluated": 0}


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v if v is not None else "").strip().lower())


def _sc_number(v):
    m = re.search(r"(\d+\.\d+(?:\.\d+)?)", str(v if v is not None else ""))
    return m.group(1) if m else None


def _normalize_conf(v):
    t = _norm(v)
    if not t:
        return None
    if "not applicable" in t or t in ("n/a", "na"):
        return "Not Applicable"
    if "not evaluated" in t or "not tested" in t or "not assessed" in t:
        return "Not Evaluated"
    if "partial" in t:
        return "Partially Supports"
    if ("does not" in t or "doesn't" in t or "not support" in t
            or t.startswith("fail") or "unsupported" in t):
        return "Does Not Support"
    if t.startswith("support") or t.startswith("pass") or t == "supported":
        return "Supports"
    return None


def _conf_from_severity(v):
    t = _norm(v)
    if any(k in t for k in ("critical", "serious", "high", "blocker", "p1", "p2")):
        return "Does Not Support"
    if any(k in t for k in ("moderate", "minor", "low", "medium", "p3", "p4")):
        return "Partially Supports"
    return None


def _worse(a, b):
    if a is None:
        return b
    if b is None:
        return a
    return a if _CONF_RANK[a] >= _CONF_RANK[b] else b


def _versions_in(text) -> set:
    t = _norm(text)
    t2 = t.replace(" ", "").replace("-", "")
    found = set()
    if "2.2" in t or "wcag22" in t2:
        found.add("2.2")
    if "2.1" in t or "wcag21" in t2:
        found.add("2.1")
    if "2.0" in t or "wcag20" in t2 or (("wcag2" in t2) and "wcag21" not in t2 and "wcag22" not in t2):
        found.add("2.0")
    return found


def _read_sheets(filename: str, raw: bytes):
    """Return a list of (sheet_name, rows) from xlsx/xls/csv (all sheets)."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv"):
        import csv as _csv
        import io as _io
        delim = "\t" if name.endswith(".tsv") else ","
        text = raw.decode("utf-8-sig", errors="replace")
        return [("CSV", [list(r) for r in _csv.reader(_io.StringIO(text), delimiter=delim)])]
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        out.append((ws.title, [list(r) for r in ws.iter_rows(values_only=True)]))
    return out


def _find_header(rows):
    """Pick the row that looks most like a header (most known keywords)."""
    keys = ("wcag version", "wcag reference", "conformance", "severity", "impact",
            "level", "status", "result", "criteria", "success criteria", "reference",
            "version")
    best_i, best_score = 0, -1
    for i, row in enumerate(rows[:40]):
        cells = [_norm(c) for c in row]
        score = sum(1 for c in cells if any(k in c for k in keys))
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def _col(headers, *names):
    for i, h in enumerate(headers):          # exact match first
        if h in names:
            return i
    for i, h in enumerate(headers):          # then contains
        if any(n in h for n in names):
            return i
    return None


_SC_RE = re.compile(r"\b\d+\.\d+\.\d+\b")                       # 1.1.1 (WCAG SC)
_VER_RE = re.compile(r"wcag\s*(?:2|20|21|22)\s*a{0,3}\b|\b2\.[012]\b", re.I)
_SEV_RE = re.compile(r"\b(critical|serious|moderate|minor|high|medium|low|blocker|p[1-4])\b", re.I)
_TAG_RE = re.compile(r"wcag\s*(2|20|21|22)\s*(a{1,3})\b", re.I)
_VER_MAP = {"2": "2.0", "20": "2.0", "21": "2.1", "22": "2.2"}
_LVL_RANK = {"A": 1, "AA": 2, "AAA": 3}


def _argmax(scores):
    if scores and max(scores) > 0:
        return max(range(len(scores)), key=lambda i: scores[i])
    return None


def _pick_columns(rows):
    """Locate reference / version / severity / status columns for one sheet,
    using header names first and column *content* as the reliable fallback."""
    hi = _find_header(rows)
    headers = [_norm(c) for c in rows[hi]] if hi < len(rows) else []
    ref_c = _col(headers, "wcag reference", "reference", "success criteria",
                 "wcag sc", "wcag criteria", "criterion", "criteria", "sc")
    ver_c = _col(headers, "wcag version", "version", "standard", "level / standard")
    status_c = _col(headers, "conformance level", "conformance", "status", "result", "supports")
    sev_c = _col(headers, "severity", "impact", "priority")

    data = rows                     # scan ALL rows: header rows won't match the
    ncol = max((len(r) for r in rows), default=0)   # SC / version / severity patterns
    refS = [0] * ncol; verS = [0] * ncol; sevS = [0] * ncol
    for r in data:
        for ci in range(min(len(r), ncol)):
            t = str(r[ci]) if r[ci] is not None else ""
            if not t:
                continue
            if _SC_RE.search(t):
                refS[ci] += 1
            if _VER_RE.search(t):
                verS[ci] += 1
            if _SEV_RE.search(t):
                sevS[ci] += 1
    # content detection wins when it has a clear signal
    if _argmax(refS) is not None:
        ref_c = _argmax(refS)
    if _argmax(verS) is not None:
        ver_c = _argmax(verS)
    if sev_c is None and _argmax(sevS) is not None:
        sev_c = _argmax(sevS)
    coverage = (max(refS) if refS else 0)
    return {"hi": hi, "ref": ref_c, "ver": ver_c, "status": status_c,
            "sev": sev_c, "coverage": coverage}


def _parse_versions(cell):
    """From a version/standard cell return (versions:set, max_level_rank:int)."""
    t = str(cell) if cell is not None else ""
    vers, rank = set(), 0
    matched = False
    for m in _TAG_RE.finditer(t):
        vers.add(_VER_MAP[m.group(1)])
        rank = max(rank, _LVL_RANK[m.group(2).upper()])
        matched = True
    if not matched:
        vers |= _versions_in(t)
    return vers, rank


# --------------------------------------------------------------------------
# Input-file validation
# The VPAT editor auto-fills the report from a SINGLE audit datasheet — the
# "placeholder 2" / All-Issues format (columns A–W of Template_WCAG_Audit.xlsx).
# That one sheet has to carry every column the delivery template reads to build
# its three sheets (see services.feature_service._fill_delivery):
#   • Sheet 2 "All Issues"      — the per-issue rows themselves
#   • Sheet 1 "Audit Summary"   — severity counts (Priority), 2.1/2.2 split
#                                 (WCAG Ref) and the fix breakdown (Fix Owner)
#   • Sheet 3 "Component Health"— grouped per Component
# Anything that is filled in by hand afterwards (the component-health rows, the
# scan-depth / AT-testing summary lines, etc.) is NOT needed in the input and is
# ignored here. If any required column is missing we refuse to open the file.
# Each entry is (human label, accepted header aliases matched as a
# case-insensitive substring of a header cell).
# --------------------------------------------------------------------------
_REQUIRED_AUDIT_COLS = [
    ("Priority",       ["priority"]),
    ("WCAG Ver",       ["wcag ver", "wcag version", "standard"]),
    ("WCAG Ref",       ["wcag ref", "wcag sc", "wcag reference",
                        "success criteria", "criterion"]),
    ("Component",      ["component"]),
    ("Issue",          ["issue"]),
    ("Issue Category", ["issue category"]),
    ("Fix Owner",      ["fix owner", "owner"]),
]


def _header_has(headers, aliases):
    """True when any alias is a substring of any (already-normalised) header."""
    return any(any(a in h for h in headers) for a in aliases)


def _missing_required_cols(sheets):
    """Scan every sheet's top rows for the header that satisfies the most
    required columns; return the labels still missing from that best header.

    Scanning a band of rows (rather than one best-guess header) keeps a
    populated sheet — whose data rows can out-score the real header — correct.
    An empty list means the file carries all required columns.
    """
    best_missing = None
    for _name, rows in sheets:
        rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        for row in rows[:15]:
            headers = [_norm(c) for c in row]
            missing = [label for label, aliases in _REQUIRED_AUDIT_COLS
                       if not _header_has(headers, aliases)]
            if best_missing is None or len(missing) < len(best_missing):
                best_missing = missing
                if not missing:
                    return []
    return best_missing if best_missing is not None else [
        label for label, _ in _REQUIRED_AUDIT_COLS]


@vpat_editor_bp.route("/api/vpat-editor/analyze", methods=["POST"])
def analyze():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No file selected."}), 400
    try:
        sheets = _read_sheets(f.filename, f.read())
    except Exception as exc:  # noqa: BLE001
        log.exception("VPAT analyse: read failed")
        return jsonify({"ok": False, "error": f"Could not read the file: {exc}"}), 400

    # The VPAT editor opens a single audit datasheet that carries every column
    # needed to fill the three delivery sheets. Reject anything else.
    missing = _missing_required_cols(sheets)
    if missing:
        return jsonify({"ok": False, "error": (
            "This file cannot be opened. The VPAT editor needs a single audit "
            "datasheet (the placeholder-2 / All-Issues format) that contains "
            "every column required to fill the three delivery sheets "
            "(Audit Summary, All Issues, Component Health). Missing column(s): "
            + ", ".join(missing) + ". Columns that are filled in manually "
            "afterwards (e.g. component-health rows) are not required.")}), 400

    # Choose the sheet whose columns carry the most WCAG references.
    best = None
    for name, rows in sheets:
        rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        if not rows:
            continue
        picked = _pick_columns(rows)
        if best is None or picked["coverage"] > best[2]["coverage"]:
            best = (name, rows, picked)
    if best is None:
        return jsonify({"ok": False, "error": "The file has no data."}), 400

    sheet_name, rows, cols = best
    hi, ref_c, ver_c, status_c, sev_c = (cols["hi"], cols["ref"], cols["ver"],
                                         cols["status"], cols["sev"])
    if ref_c is None and ver_c is None:
        sample = [str(c) for c in rows[hi]][:12] if hi < len(rows) else []
        return jsonify({"ok": False, "error": (
            "Couldn't locate WCAG data. Looked for a column with success-criterion "
            "numbers (e.g. 1.1.1) and a WCAG version/standard column, across every "
            "sheet. Checked sheet '" + sheet_name + "'. First row seen: "
            + ", ".join(sample)[:280])}), 400

    versions, tag_rank = set(), 0
    per_sc = {}
    finding_rows = 0
    for row in rows:
        if ver_c is not None and ver_c < len(row):
            v, rk = _parse_versions(row[ver_c])
            versions |= v
            tag_rank = max(tag_rank, rk)
        if ref_c is None or ref_c >= len(row):
            continue
        sc = _sc_number(row[ref_c])
        if not sc:
            continue
        finding_rows += 1
        rec = per_sc.setdefault(sc, {"explicit": None, "sev": None, "count": 0})
        rec["count"] += 1
        if status_c is not None and status_c < len(row):
            rec["explicit"] = _worse(rec["explicit"], _normalize_conf(row[status_c]))
        if sev_c is not None and sev_c < len(row):
            rec["sev"] = _worse(rec["sev"], _conf_from_severity(row[sev_c]))

    criteria_status = {}
    criteria_counts = {sc: rec["count"] for sc, rec in per_sc.items()}
    for sc, rec in per_sc.items():
        if rec["explicit"] is not None:
            conf = rec["explicit"]
        elif rec["sev"] is not None:
            conf = rec["sev"]
        else:
            conf = "Does Not Support" if rec["count"] >= 3 else "Partially Supports"
        criteria_status[sc] = conf

    tag_level = {1: "A", 2: "AA", 3: "AAA"}.get(tag_rank)

    # Build a small preview of the uploaded file so the user can cross-check.
    PREVIEW_ROWS, PREVIEW_COLS, CELL_CAP = 200, 30, 300
    ncol = min(max((len(r) for r in rows), default=0), PREVIEW_COLS)
    preview_rows = []
    for r in rows[:PREVIEW_ROWS]:
        cells = []
        for ci in range(ncol):
            v = r[ci] if ci < len(r) else None
            s = "" if v is None else str(v)
            cells.append(s[:CELL_CAP] + ("…" if len(s) > CELL_CAP else ""))
        preview_rows.append(cells)
    preview = {
        "filename": f.filename,
        "sheet": sheet_name,
        "header_index": hi,
        "rows": preview_rows,
        "total_rows": len(rows),
        "ncol": ncol,
        "detected": {"reference": ref_c, "version": ver_c,
                     "severity": sev_c, "status": status_c},
    }

    return jsonify({
        "ok": True,
        "sheet": sheet_name,
        "versions": sorted(versions),
        "tag_level": tag_level,
        "criteria_status": criteria_status,
        "criteria_counts": criteria_counts,
        "had_status_col": status_c is not None,
        "had_severity_col": sev_c is not None,
        "rows": finding_rows,
        "criteria_count": len(per_sc),
        "preview": preview,
    })


@vpat_editor_bp.route("/vpat-editor")
def page():
    # Start from the JBL report (keeps Standards, Terms, criteria & Legal as
    # sensible defaults) but BLANK the Overview metadata so the user types it.
    data = default_data()
    for k in ("title", "subtitle", "product", "report_date",
              "description", "contact", "notes", "eval_methods"):
        data[k] = ""
    resp = make_response(render_template(
        "vpat_editor.html", data=data, slugs=WCAG_SLUGS, build=VPAT_BUILD))
    # This page embeds its CSS/JS inline, so prevent the browser from serving a
    # cached copy of the HTML (otherwise layout changes won't appear).
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@vpat_editor_bp.route("/api/vpat-editor/default")
def get_default():
    return jsonify({"ok": True, "data": default_data()})


@vpat_editor_bp.route("/api/vpat-editor/export-pdf", methods=["POST"])
def export():
    data = request.get_json(force=True, silent=True) or {}
    if not data.get("level_a") and not data.get("title"):
        data = default_data()
    product = re.sub(r"\s+", "_", str(data.get("product", "report")).strip()) or "report"
    stem = file_utils.safe_filename(
        f"VPAT_{product}_{datetime.datetime.now():%Y%m%d_%H%M%S}").replace(" ", "_")
    out = (config.export_dir / f"{stem}.pdf")
    try:
        export_pdf(data, str(out))
    except ImportError:
        return jsonify({"ok": False, "error": (
            "PDF export needs the 'reportlab' package, which isn't installed. "
            "Run:  pip install reportlab  (or  pip install -r requirements.txt)  "
            "then restart the app.")}), 400
    except Exception as exc:  # noqa: BLE001
        log.exception("VPAT PDF export failed")
        return jsonify({"ok": False, "error": f"PDF export failed: {exc}"}), 400
    return jsonify({"ok": True, "file": out.name,
                    "url": f"/download/exports/{out.name}"})


@vpat_editor_bp.route("/api/vpat-editor/save-draft", methods=["POST"])
def save_draft():
    """Persist the current report as a JSON draft and return a download link."""
    data = request.get_json(force=True, silent=True) or {}
    product = re.sub(r"\s+", "_", str(data.get("product", "report")).strip()) or "report"
    stem = file_utils.safe_filename(
        f"VPAT_draft_{product}_{datetime.datetime.now():%Y%m%d_%H%M%S}").replace(" ", "_")
    out = (config.export_dir / f"{stem}.json")
    try:
        out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        log.exception("VPAT save-draft failed")
        return jsonify({"ok": False, "error": f"Save failed: {exc}"}), 400
    return jsonify({"ok": True, "file": out.name,
                    "url": f"/download/exports/{out.name}"})


@vpat_editor_bp.route("/api/vpat-editor/open-draft", methods=["POST"])
def open_draft():
    """Parse an uploaded JSON draft and return its data to repopulate the form."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "No draft file selected."}), 400
    try:
        data = json.loads(f.read().decode("utf-8"))
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"Invalid draft JSON: {exc}"}), 400
    return jsonify({"ok": True, "data": data})


# --------------------------------------------------------------------------
# Server-side named drafts: "Save As" + "Recent files" (with date/time)
# --------------------------------------------------------------------------
def _drafts_dir():
    d = config.data_dir / "vpat_drafts"
    d.mkdir(parents=True, exist_ok=True)
    return d


@vpat_editor_bp.route("/api/vpat-editor/save-as", methods=["POST"])
def save_as():
    """Save the current report under a user-chosen name (server-side)."""
    payload = request.get_json(force=True, silent=True) or {}
    name = str(payload.get("name", "")).strip()
    data = payload.get("data") or {}
    if not name:
        return jsonify({"ok": False, "error": "Please enter a name for the report."}), 400
    stem = file_utils.safe_filename(name).replace(" ", "_") or "vpat_report"
    if stem.lower().endswith(".json"):
        stem = stem[:-5]
    out = _drafts_dir() / f"{stem}.json"
    try:
        out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        log.exception("VPAT save-as failed")
        return jsonify({"ok": False, "error": f"Save failed: {exc}"}), 400
    return jsonify({"ok": True, "name": out.stem, "file": out.name})


@vpat_editor_bp.route("/api/vpat-editor/recent")
def recent():
    """List server-side saved reports, newest first, with date/time."""
    items = []
    for p in _drafts_dir().glob("*.json"):
        try:
            ts = p.stat().st_mtime
        except OSError:
            continue
        items.append({
            "name": p.stem,
            "file": p.name,
            "modified": datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S"),
            "ts": ts,
        })
    items.sort(key=lambda x: x["ts"], reverse=True)
    return jsonify({"ok": True, "items": items[:50]})


@vpat_editor_bp.route("/api/vpat-editor/load-draft/<path:name>")
def load_draft(name):
    """Return the data for a previously saved (server-side) report."""
    stem = file_utils.safe_filename(name).replace(" ", "_")
    if stem.lower().endswith(".json"):
        stem = stem[:-5]
    p = _drafts_dir() / f"{stem}.json"
    if not p.exists():
        return jsonify({"ok": False, "error": "That saved report no longer exists."}), 404
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return jsonify({"ok": False, "error": f"Could not read report: {exc}"}), 400
    return jsonify({"ok": True, "data": data, "name": p.stem})
