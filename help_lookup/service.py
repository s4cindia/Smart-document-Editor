"""Help-lookup data + search — fully self-contained.

This module powers the delivery editor's "WCAG & Axe Lookup" Help dialog and is
intentionally decoupled from the rest of the app so it can be evolved on its own:

  * It owns its own data-file paths (no imports from feature_service).
  * It owns its own loaders/search for both the WCAG tag file and the axe
    rule-id mapping spreadsheet.

Data files (bundled under data/templates/):
  * wcag_tags.txt          — "id|name|level" per line (WCAG criteria tab)
  * Rule_ID_Mapping.xlsx   — Rule ID / Description / Help / Help URL / Tags /
                             ACT IDs (the "Rule ID" tab)

To change what Help shows, edit those two files and/or this module — nothing
else in the codebase needs touching.
"""
from __future__ import annotations

import re
from pathlib import Path

from config import config

TEMPLATE_DIR = config.data_dir / "templates"
WCAG_TAGS_FILE = TEMPLATE_DIR / "wcag_tags.txt"
RULE_ID_FILE = TEMPLATE_DIR / "Rule_ID_Mapping.xlsx"


# --------------------------------------------------------------------------- #
# WCAG criteria  (wcag_tags.txt)
# --------------------------------------------------------------------------- #
def load_wcag_tags(path: Path = WCAG_TAGS_FILE) -> dict[str, tuple[str, str]]:
    """Parse wcag_tags.txt -> {id: (name, level)}. Returns {} if missing."""
    out: dict[str, tuple[str, str]] = {}
    if not path or not Path(path).exists():
        return out
    try:
        for line in Path(path).read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or "|" not in line:
                continue
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                out[parts[0]] = (parts[1], parts[2])
            elif len(parts) == 2:
                out[parts[0]] = (parts[1], "")
    except Exception:  # noqa: BLE001
        return out
    return out


def search_wcag(q: str = "") -> dict:
    """Search WCAG criteria by ref / name / level. Sorted by criterion number."""
    q = (q or "").strip().lower()
    tags = load_wcag_tags()
    rows = [{"ref": k, "name": v[0], "level": v[1]} for k, v in tags.items()]
    if q:
        rows = [r for r in rows
                if q in r["ref"].lower() or q in r["name"].lower()
                or q in r["level"].lower()]

    def _key(r):
        return [int(p) if p.isdigit() else 0 for p in r["ref"].split(".")]

    rows.sort(key=_key)
    return {"rows": rows, "count": len(rows), "total": len(tags)}


# --------------------------------------------------------------------------- #
# axe rules  (Rule_ID_Mapping.xlsx)
# --------------------------------------------------------------------------- #
_CRIT_RE = re.compile(r"^wcag(\d{3,4})$", re.I)
_LEVEL_RE = re.compile(r"^wcag2[12]?(a{1,3})$", re.I)
_LEVEL_RANK = {"A": 1, "AA": 2, "AAA": 3}


def _wcag_from_tags(tags) -> tuple[list[str], str, bool]:
    """Derive WCAG criteria (e.g. 1.4.3), level (A/AA/AAA) and best-practice
    flag from an axe rule's tag list."""
    crits, level = [], ""
    for t in tags or []:
        m = _CRIT_RE.match(str(t))
        if m:
            d = m.group(1)
            crit = f"{d[0]}.{d[1]}.{d[2:]}"
            if crit not in crits:
                crits.append(crit)
        m2 = _LEVEL_RE.match(str(t))
        if m2:
            lvl = m2.group(1).upper()
            if _LEVEL_RANK.get(lvl, 0) > _LEVEL_RANK.get(level, 0):
                level = lvl
    return crits, level, ("best-practice" in (tags or []))


def _split(v) -> list[str]:
    if v is None:
        return []
    return [p.strip() for p in str(v).replace(",", ";").split(";") if p.strip()]


def load_axe_rules(path: Path = RULE_ID_FILE) -> list[dict]:
    """Load the Rule ID mapping spreadsheet -> list of rule dicts.

    Columns (header row, order-tolerant): Rule ID, Description, Help, Help URL,
    Tags (';'-separated), ACT IDs (';'-separated). Returns [] if unreadable."""
    if not path or not Path(path).exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:  # noqa: BLE001
        return []
    if not rows:
        return []

    hdr = {str(h).strip().lower(): i for i, h in enumerate(rows[0]) if h is not None}

    def _col(*names):
        for n in names:
            if n in hdr:
                return hdr[n]
        return None

    i_id = _col("rule id", "ruleid")
    i_desc = _col("description")
    i_help = _col("help")
    i_url = _col("help url", "helpurl")
    i_tags = _col("tags")
    i_act = _col("act ids", "actids")
    out = []
    for r in rows[1:]:
        def cell(i):
            return r[i] if (i is not None and i < len(r)) else None

        rid = cell(i_id)
        if rid is None or not str(rid).strip():
            continue
        out.append({
            "ruleId": str(rid).strip(),
            "description": str(cell(i_desc) or "").strip(),
            "help": str(cell(i_help) or "").strip(),
            "helpUrl": str(cell(i_url) or "").strip(),
            "tags": _split(cell(i_tags)),
            "actIds": _split(cell(i_act)),
        })
    return out


def search_axe(q: str = "") -> dict:
    """Token-based search of the axe catalogue by ruleId / description / help /
    tag / derived WCAG criterion. Each row carries ready-to-display fields."""
    tokens = (q or "").strip().lower().split()
    rules = load_axe_rules()
    out = []
    for r in rules:
        rid, desc, helptxt = r["ruleId"], r["description"], r["help"]
        tags = r["tags"]
        crits, level, best = _wcag_from_tags(tags)
        hay = " ".join([rid, desc, helptxt, " ".join(tags), " ".join(crits)]).lower()
        if tokens and not all(tok in hay for tok in tokens):
            continue
        out.append({
            "ruleId": rid, "description": desc, "help": helptxt,
            "helpUrl": r["helpUrl"], "tags": tags,
            "wcag": ", ".join(crits),
            "level": ("Best practice" if (best and not level) else level),
        })
    out.sort(key=lambda x: x["ruleId"])
    return {"rows": out, "count": len(out), "total": len(rules)}
